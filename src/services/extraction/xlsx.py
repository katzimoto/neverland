"""XLSX text extractor."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


class XlsxExtractor:
    """Extract text from Excel .xlsx files using openpyxl."""

    def extract(self, path: Path) -> str:
        """Return concatenated text from all cells."""
        try:
            wb = load_workbook(str(path))
            texts: list[str] = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            texts.append(str(cell.value))
            return "\n".join(texts)
        except (OSError, KeyError):
            return ""
